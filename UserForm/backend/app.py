from flask import Flask, request, jsonify
from flask_cors import CORS 
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime

app = Flask(__name__)
CORS(app)
# Define the path to the Excel file
EXCEL_FILE = 'user_data.xlsx'

@app.route('/', methods=['GET'])
def index():
    return "Flask is running!"

@app.route('/save', methods=['POST'])
def save_data():
    data = request.get_json()
    name = data.get("name")
    email = data.get("email")
    date_str = datetime.now().strftime('%Y-%m-%d')
    if not name or not email:
        return jsonify({"error": "Name and email are required"}), 400

    # Check if the Excel file exists, if not create it
    if not os.path.exists(EXCEL_FILE):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Name", "Email", "CreatedOn"])
        workbook.save(EXCEL_FILE)
    
    # Load the existing workbook and append the new data
    workbook = load_workbook(EXCEL_FILE)

    if 'Sheet' not in workbook.sheetnames:
        workbook.create_sheet('Sheet')
        workbook['Sheet'].append(["Name", "Email", "CreatedOn"])

    sheet = workbook['Sheet']
    created_on = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sheet.append([name, email, created_on])

    if date_str not in workbook.sheetnames:
        workbook.create_sheet(date_str)
        workbook[date_str].append(["Name", "Email", "CreatedOn"])
    date_sheet = workbook[date_str]
    date_sheet.append([name, email, created_on])


    workbook.save(EXCEL_FILE)   
    return jsonify({"message": "Data saved successfully"}), 200

    if __name__ == '__main__':
        app.run(debug=True)